import asyncio
import sqlite3
import os
import re
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import Command
from aiogram.types import FSInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from apscheduler.schedulers.asyncio import AsyncIOScheduler

# --- 1. НАСТРОЙКИ ---
TOKEN = '8322200956:AAFiFBqnXd91dH04BY8d0WnwJFE_yZEAo9w'
ADMIN_ID = 1992812428
bot = Bot(token=TOKEN)
dp = Dispatcher()

# --- 2. БАЗА ДАННЫХ ---
def init_db():
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT UNIQUE,
            expiry_date TEXT,
            has_ecp INTEGER DEFAULT 1
        )
    ''')
    conn.commit()
    conn.close()

def add_or_update_user(full_name: str, date_str: str = None, has_ecp: int = 1):
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    if date_str:
        cur.execute('REPLACE INTO users (full_name, expiry_date, has_ecp) VALUES (?, ?, ?)', 
                   (full_name, date_str, has_ecp))
    else:
        cur.execute('REPLACE INTO users (full_name, expiry_date, has_ecp) VALUES (?, ?, ?)', 
                   (full_name, None, has_ecp))
    conn.commit()
    conn.close()

def remove_user(full_name: str):
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    cur.execute('DELETE FROM users WHERE full_name = ?', (full_name,))
    deleted = cur.rowcount > 0
    conn.commit()
    conn.close()
    return deleted

def find_users_by_name(query: str):
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    cur.execute('SELECT full_name, expiry_date, has_ecp FROM users WHERE full_name LIKE ? ORDER BY expiry_date ASC', 
                (f'%{query}%',))
    results = cur.fetchall()
    conn.close()
    return results

def get_all_users():
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    cur.execute('SELECT full_name, expiry_date, has_ecp FROM users ORDER BY expiry_date ASC')
    data = cur.fetchall()
    conn.close()
    return data

def get_expiring_soon():
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    today = datetime.now().date()
    target_date = today + timedelta(days=60)  # ИЗМЕНЕНО: 30 -> 60 дней
    
    cur.execute('''
        SELECT full_name, expiry_date, has_ecp FROM users 
        WHERE has_ecp = 1 AND expiry_date <= ?
        ORDER BY expiry_date ASC
    ''', (target_date,))
    expiring_ecp = cur.fetchall()
    
    cur.execute('''
        SELECT full_name, expiry_date, has_ecp FROM users 
        WHERE has_ecp = 0
        ORDER BY full_name ASC
    ''')
    no_ecp = cur.fetchall()
    
    conn.close()
    
    result = []
    for name, exp_date_str, has_ecp in expiring_ecp:
        exp_date = datetime.strptime(exp_date_str, '%Y-%m-%d').date()
        days_left = (exp_date - today).days
        if days_left >= 0:
            result.append((name, exp_date_str, days_left, has_ecp))
    
    for name, exp_date_str, has_ecp in no_ecp:
        result.append((name, None, None, has_ecp))
    
    return result

# --- 3. АРХИВАЦИЯ (каждую пятницу в 19:00) ---
async def create_archive():
    """Создает архивный файл и отправляет в чат"""
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Создание еженедельного архива...")
    
    data = get_all_users()
    if not data:
        print("  База данных пуста, архив не создан")
        await bot.send_message(
            ADMIN_ID, 
            "📭 *Архив не создан*\n\nБаза данных пуста. Добавьте сотрудников через /запись или /добавить",
            parse_mode="Markdown"
        )
        return
    
    date_str = datetime.now().strftime('%Y-%m-%d_%H-%M')
    file_name = f"archive_ECP_{date_str}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Архив ЭЦП"
    
    headers = ["№", "ФИО сотрудника", "Статус", "Срок действия ЭЦП", "Осталось дней на момент архива"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    today = datetime.now().date()
    
    for idx, (name, date_str, has_ecp) in enumerate(data, start=1):
        if has_ecp == 0:
            status = "❌ НЕТ ЭЦП"
            formatted_date = "Не выпущена"
            days_left_text = "Требуется выпуск!"
        else:
            status = "✅ Есть ЭЦП"
            exp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            days_left = (exp_date - today).days
            formatted_date = exp_date.strftime('%d.%m.%Y')
            days_left_text = f"{days_left} дней"
        
        ws.append([idx, name, status, formatted_date, days_left_text])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    wb.save(file_name)
    print(f"  ✅ Архив создан: {file_name}")
    
    document = FSInputFile(file_name)
    await bot.send_document(
        ADMIN_ID,
        document,
        caption=f"📦 *Еженедельный архив ЭЦП*\n\n"
                f"📅 Дата: {datetime.now().strftime('%d.%m.%Y')}\n"
                f"⏰ Время: {datetime.now().strftime('%H:%M')}\n"
                f"👥 Всего сотрудников: {len(data)}\n\n"
                f"📌 *Архив создан автоматически*\n"
                f"_Для актуальных данных используйте /выгрузка_",
        parse_mode="Markdown"
    )
    
    os.remove(file_name)
    print(f"  🗑️ Временный файл удален: {file_name}")

# --- 4. ЕЖЕДНЕВНАЯ РАССЫЛКА (в 8:00) ---
async def send_daily_report():
    print(f"⏰ Ежедневная проверка в {datetime.now()}")
    expiring = get_expiring_soon()
    
    if not expiring:
        await bot.send_message(ADMIN_ID, "✅ *ПОРЯДОК!* \n✨ У всех всё хорошо! ☕️😊", parse_mode="Markdown")
        return
    
    message = "📋 *ЕЖЕДНЕВНЫЙ ОТЧЕТ ПО ЭЦП* 📋\n\n"
    message += f"📅 *Период:* следующие 60 дней\n\n"  # ИЗМЕНЕНО
    
    no_ecp_list = [item for item in expiring if item[3] == 0]
    expiring_list = [item for item in expiring if item[3] == 1]
    
    if no_ecp_list:
        message += "🆕 *СОТРУДНИКИ БЕЗ ЭЦП:*\n"
        for name, _, _, _ in no_ecp_list:
            message += f"   👤 *{name}* — 🔴 *ТРЕБУЕТСЯ ВЫПУСК ЭЦП!*\n"
        message += "\n"
    
    if expiring_list:
        message += "⚠️ *ИСТЕКАЕТ СРОК ЭЦП:*\n"
        for name, date_str, days_left, _ in expiring_list:
            formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d.%m.%Y')
            
            if days_left <= 7:
                emoji = "🔴🔥⚠️"
            elif days_left <= 14:
                emoji = "🟠⚠️"
            else:
                emoji = "🟡❗️"
            
            if days_left % 10 == 1 and days_left % 100 != 11:
                day_word = "день"
            elif 2 <= days_left % 10 <= 4 and (days_left % 100 < 10 or days_left % 100 >= 20):
                day_word = "дня"
            else:
                day_word = "дней"
                
            message += f"{emoji} *{name}*\n   📅 {formatted_date} — осталось *{days_left}* {day_word}\n\n"
    
    await bot.send_message(ADMIN_ID, message, parse_mode="Markdown")

# --- 5. КОМАНДА /дай ---
@dp.message(Command("дай"))
async def give_report(message: types.Message):
    await message.answer("📊 *Формирую актуальный отчет...*", parse_mode="Markdown")
    
    expiring = get_expiring_soon()
    
    if not expiring:
        await message.answer("✅ *ВСЁ В ПОРЯДКЕ!* \n✨ Ни у кого нет проблем с ЭЦП! 🎉", parse_mode="Markdown")
        return
    
    report = "📋 *АКТУАЛЬНЫЙ ОТЧЕТ ПО ЭЦП* 📋\n\n"
    report += f"📅 *Период:* следующие 60 дней\n\n"  # ИЗМЕНЕНО
    
    no_ecp_list = [item for item in expiring if item[3] == 0]
    expiring_list = [item for item in expiring if item[3] == 1]
    
    if no_ecp_list:
        report += "🆕 *НУЖНО ВЫПУСТИТЬ ЭЦП:*\n"
        for name, _, _, _ in no_ecp_list:
            report += f"   👤 *{name}* — 🔴 *ТРЕБУЕТСЯ ВЫПУСК ЭЦП!*\n"
        report += "\n"
    
    if expiring_list:
        report += "⚠️ *ИСТЕКАЕТ СРОК ЭЦП:*\n"
        for name, date_str, days_left, _ in expiring_list:
            formatted_date = datetime.strptime(date_str, '%Y-%m-%d').strftime('%d.%m.%Y')
            
            if days_left <= 7:
                emoji = "🔴🔥⚠️"
            elif days_left <= 14:
                emoji = "🟠⚠️"
            else:
                emoji = "🟡❗️"
            
            if days_left % 10 == 1 and days_left % 100 != 11:
                day_word = "день"
            elif 2 <= days_left % 10 <= 4 and (days_left % 100 < 10 or days_left % 100 >= 20):
                day_word = "дня"
            else:
                day_word = "дней"
                
            report += f"{emoji} *{name}*\n   📅 {formatted_date} — осталось *{days_left}* {day_word}\n\n"
    
    await message.answer(report, parse_mode="Markdown")

# --- 6. КОМАНДА /удалить ---
@dp.message(Command("удалить"))
async def delete_employee(message: types.Message):
    text = message.text.replace('/удалить', '').strip()
    
    if not text:
        await message.answer(
            "❌ *Укажите ФИО сотрудника!*\n\n"
            "Пример: `/удалить Иванов Иван Иванович`",
            parse_mode="Markdown"
        )
        return
    
    full_name = text.strip()
    results = find_users_by_name(full_name)
    exact_match = any(name.lower() == full_name.lower() for name, _, _ in results)
    
    if not exact_match:
        await message.answer(f"❌ *Сотрудник {full_name} не найден в базе!*", parse_mode="Markdown")
        return
    
    if remove_user(full_name):
        await message.answer(f"👋 *Сотрудник удален из базы!*\n\n👤 {full_name}\n✅ Больше не отслеживается.", parse_mode="Markdown")
    else:
        await message.answer(f"❌ Ошибка при удалении", parse_mode="Markdown")

# --- 7. КОМАНДА /добавить ---
@dp.message(Command("добавить"))
async def add_employee(message: types.Message):
    text = message.text.replace('/добавить', '').strip()
    
    if not text:
        await message.answer(
            "❌ *Укажите ФИО сотрудника!*\n\n"
            "✅ `/добавить Петров Петр Петрович` — новый сотрудник БЕЗ ЭЦП\n"
            "✅ `/запись Петров Петр Петрович 13.07.2026` — сотрудник С ЭЦП",
            parse_mode="Markdown"
        )
        return
    
    if re.search(r'\d{2}\.\d{2}\.\d{4}', text):
        await message.answer(
            "❌ *Ошибка!*\n\n"
            "Команда `/добавить` НЕ используется с датой.\n\n"
            "📌 *Правильно:*\n"
            "• Без ЭЦП: `/добавить Иванов Иван Иванович`\n"
            "• С ЭЦП: `/запись Иванов Иван Иванович 13.07.2026`",
            parse_mode="Markdown"
        )
        return
    
    full_name = text.strip()
    
    if len(full_name.split()) < 2:
        await message.answer("❌ *Укажите полное ФИО!*", parse_mode="Markdown")
        return
    
    results = find_users_by_name(full_name)
    for name, _, has_ecp in results:
        if name.lower() == full_name.lower():
            if has_ecp == 1:
                await message.answer(
                    f"⚠️ *{full_name} уже есть в базе с ЭЦП!*\n\n"
                    f"Используйте `/запись {full_name} ДД.ММ.ГГГГ` для обновления",
                    parse_mode="Markdown"
                )
            else:
                await message.answer(f"⚠️ *{full_name} уже есть в базе* (нет ЭЦП)", parse_mode="Markdown")
            return
    
    add_or_update_user(full_name, date_str=None, has_ecp=0)
    await message.answer(
        f"🎉 *Новый сотрудник добавлен!*\n\n"
        f"👤 *{full_name}*\n"
        f"🆕 Статус: *нет ЭЦП*\n\n"
        f"Когда сотрудник получит ЭЦП, используйте:\n"
        f"`/запись {full_name} ДД.ММ.ГГГГ`",
        parse_mode="Markdown"
    )

# --- 8. КОМАНДА /выгрузка ---
@dp.message(Command("выгрузка"))
async def export_excel(message: types.Message):
    await message.answer("📊 *Формирую Excel-таблицу...*", parse_mode="Markdown")
    
    data = get_all_users()
    if not data:
        await message.answer("📭 База данных пуста.", parse_mode="Markdown")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сроки ЭЦП"
    
    headers = ["№", "ФИО сотрудника", "Статус", "Срок действия ЭЦП", "Осталось дней"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    today = datetime.now().date()
    
    for idx, (name, date_str, has_ecp) in enumerate(data, start=1):
        if has_ecp == 0:
            status = "❌ НЕТ ЭЦП"
            formatted_date = "Не выпущена"
            days_left_text = "Требуется выпуск!"
            fill_color = None
        else:
            status = "✅ Есть ЭЦП"
            exp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            days_left = (exp_date - today).days
            formatted_date = exp_date.strftime('%d.%m.%Y')
            days_left_text = f"{days_left} дней"
            
            if days_left <= 7:
                fill_color = "FF6B6B"
            elif days_left <= 14:
                fill_color = "FFA500"
            elif days_left <= 60:  # ИЗМЕНЕНО: 30 -> 60 дней для подсветки
                fill_color = "FFD700"
            else:
                fill_color = None
        
        ws.append([idx, name, status, formatted_date, days_left_text])
        
        if fill_color:
            for cell in ws[idx + 1]:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    file_name = f"E-Certificates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(file_name)
    
    document = FSInputFile(file_name)
    await message.answer_document(
        document, 
        caption="📊 *Актуальная таблица по ЭЦП*\n\n"
                "📌 *Легенда:*\n"
                "• ❌ НЕТ ЭЦП — требуется выпуск\n"
                "• 🔴 Красный — осталось 7 дней и меньше\n"
                "• 🟠 Оранжевый — осталось 14 дней и меньше\n"
                "• 🟡 Желтый — осталось 60 дней и меньше",  # ИЗМЕНЕНО
        parse_mode="Markdown"
    )
    
    os.remove(file_name)

# --- 9. КОМАНДА /запись ---
@dp.message(Command("запись"))
async def add_record(message: types.Message):
    text = message.text.replace('/запись', '').strip()
    parts = text.split()
    
    if len(parts) < 4:
        await message.answer(
            "❌ *Формат:* `/запись Фамилия Имя Отчество ДД.ММ.ГГГГ`\n\n"
            "📌 *Пример:* `/запись Иванов Иван Иванович 13.07.2026`",
            parse_mode="Markdown"
        )
        return
    
    date_str = parts[-1]
    full_name = ' '.join(parts[:-1])
    
    if not re.match(r'\d{2}\.\d{2}\.\d{4}', date_str):
        await message.answer(
            "❌ *Неверный формат даты!*\n\n"
            "Используйте ДД.ММ.ГГГГ, например: `13.07.2026`",
            parse_mode="Markdown"
        )
        return
    
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        
        if date_obj.date() < datetime.now().date():
            await message.answer(f"⚠️ Дата {date_str} уже прошла! Укажите будущую дату.", parse_mode="Markdown")
            return
        
        db_date = date_obj.strftime('%Y-%m-%d')
        
        results = find_users_by_name(full_name)
        for name, _, has_ecp in results:
            if name.lower() == full_name.lower() and has_ecp == 0:
                await message.answer(
                    f"✏️ *Обновлено!* ✅\n\n"
                    f"👤 *{full_name}*\n"
                    f"📅 {date_str}\n\n"
                    f"🎉 Сотрудник получил ЭЦП! Теперь он в системе с датой окончания.",
                    parse_mode="Markdown"
                )
                add_or_update_user(full_name, db_date, has_ecp=1)
                return
        
        add_or_update_user(full_name, db_date, has_ecp=1)
        
        days_left = (date_obj.date() - datetime.now().date()).days
        response = f"✏️ *Записал!* ✅\n\n👤 *{full_name}*\n📅 {date_str}\n"
        
        if days_left <= 60:  # ИЗМЕНЕНО: 30 -> 60 дней
            response += f"⚠️ *Осталось {days_left} дней!* Не забудьте продлить!"
        else:
            response += f"📌 Осталось {days_left} дней"
        
        await message.answer(response, parse_mode="Markdown")
        
    except ValueError:
        await message.answer("❌ *Неверный формат даты!* Используйте ДД.ММ.ГГГГ", parse_mode="Markdown")

# --- 10. КОМАНДА /start ---
@dp.message(Command("start"))
async def start_cmd(message: types.Message):
    await message.answer(
        "👋 *Бот по учету ЭЦП*\n\n"
        "📌 *Команды:*\n"
        "• `/запись ФИО ДД.ММ.ГГГГ` — добавить/обновить ЭЦП\n"
        "• `/добавить ФИО` — добавить сотрудника (без ЭЦП)\n"
        "• `/удалить ФИО` — удалить сотрудника\n"
        "• `/дай` — отчет по проблемным (за 60 дней)\n"
        "• `/Фамилия` — найти сотрудника\n"
        "• `/выгрузка` — скачать Excel\n"
        "• `/статистика` — статистика\n\n"
        "🔔 *Ежедневно в 8:00* приходит отчет (период: 60 дней)\n"
        "📦 *Каждую пятницу в 19:00* приходит архив",
        parse_mode="Markdown"
    )

# --- 11. КОМАНДА /help ---
@dp.message(Command("help"))
async def help_cmd(message: types.Message):
    await message.answer(
        "📚 *Справка по командам:*\n\n"
        "🔹 `/запись Иванов Иван Иванович 13.07.2026` — добавить/обновить ЭЦП\n"
        "🔹 `/добавить Петров Петр Петрович` — добавить сотрудника (без ЭЦП)\n"
        "🔹 `/удалить Сидоров Сидор Сидорович` — удалить сотрудника\n"
        "🔹 `/дай` — получить актуальный отчет за 60 дней\n"
        "🔹 `/Иванов` — найти сотрудника\n"
        "🔹 `/выгрузка` — скачать Excel-таблицу\n"
        "🔹 `/статистика` — общая статистика\n\n"
        "⏰ *Ежедневно в 8:00* приходит отчет (период: 60 дней)\n"
        "📦 *Каждую пятницу в 19:00* приходит архив в формате Excel",
        parse_mode="Markdown"
    )

# --- 12. КОМАНДА /статистика ---
@dp.message(Command("статистика"))
async def stats_cmd(message: types.Message):
    conn = sqlite3.connect('certificates.db')
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM users')
    total = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM users WHERE has_ecp = 0')
    no_ecp = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM users WHERE has_ecp = 1 AND expiry_date <= date("now", "+60 days")')  # ИЗМЕНЕНО
    expiring = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM users WHERE has_ecp = 1 AND expiry_date < date("now")')
    expired = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM users WHERE has_ecp = 1 AND expiry_date > date("now", "+60 days")')  # ИЗМЕНЕНО
    valid = cur.fetchone()[0]
    conn.close()
    
    await message.answer(
        f"📊 *Статистика базы ЭЦП*\n\n"
        f"👥 *Всего сотрудников:* {total}\n"
        f"🆕 *Нет ЭЦП (нужно выпустить):* {no_ecp}\n"
        f"✅ *ЭЦП в порядке:* {valid}\n"
        f"⚠️ *Истекает в ближайшие 60 дней:* {expiring}\n"  # ИЗМЕНЕНО
        f"❗️ *Просрочено:* {expired}\n\n"
        f"💡 Используйте `/дай` для детального отчета",
        parse_mode="Markdown"
    )

# --- 13. ПОИСК ПО ИМЕНИ ---
@dp.message(F.text.startswith('/'))
async def search_user(message: types.Message):
    query = message.text[1:]
    known = ['start', 'выгрузка', 'запись', 'help', 'статистика', 'дай', 'удалить', 'добавить']
    if query.split()[0] in known:
        return
    
    results = find_users_by_name(query.strip())
    if not results:
        await message.answer(f"🤷 *Ничего не найдено* по запросу `{query}`", parse_mode="Markdown")
        return
    
    today = datetime.now().date()
    answer = f"🔎 *Найдено совпадений:* {len(results)}\n\n"
    
    for name, date_str, has_ecp in results:
        if has_ecp == 0:
            answer += f"🆕 *{name}*\n   ❌ *НЕТ ЭЦП* — требуется выпуск!\n\n"
        else:
            exp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            days = (exp_date - today).days
            formatted = exp_date.strftime('%d.%m.%Y')
            
            if days < 0:
                status = f"❗️ *ПРОСРОЧЕНА* на {abs(days)} дней!"
                emoji = "💀"
            elif days <= 7:
                status = f"🔴 Осталось *{days}* дней — КРИТИЧНО!"
                emoji = "🚨"
            elif days <= 60:  # ИЗМЕНЕНО: 30 -> 60 дней
                status = f"🟠 Осталось *{days}* дней — скоро истекает"
                emoji = "⚡️"
            else:
                status = f"✅ Осталось {days} дней"
                emoji = "📌"
                
            answer += f"{emoji} *{name}*\n   📅 {formatted} — {status}\n\n"
    
    await message.answer(answer, parse_mode="Markdown")

# --- 14. ЗАПУСК БОТА ---
async def main():
    init_db()
    
    scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
    
    scheduler.add_job(send_daily_report, "cron", hour=8, minute=0)
    scheduler.add_job(create_archive, "cron", day_of_week='fri', hour=19, minute=0)
    
    scheduler.start()
    
    print("✅ БОТ ЗАПУЩЕН!")
    print("⏰ Ежедневная рассылка: 8:00 МСК (период: 60 дней)")
    print("📦 Еженедельный архив: Пятница 19:00 МСК")
    print("🤖 Бот готов к работе!")
    
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())